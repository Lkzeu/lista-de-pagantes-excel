#!/usr/bin/env python3

from datetime import datetime
from typing import Dict

from openpyxl import load_workbook

from functions.send_messages.vencimentos_txt import get_vencimento, dia_vencimento
from functions.send_messages.verifica_se_eh_brazileiro import is_brasileiro

NUMERO = str
VENCIMENTO = str
def numbers_xl_to_txt(arquivo: str, mes: str) -> Dict[NUMERO, VENCIMENTO]:
   numeros_e_vencimentos = dict()

   wb = load_workbook(arquivo)
   ws = wb[mes]

   numero_dos_pagantes = ws['B']

   agora = datetime.now()
   hoje = agora.day

   min, max = 0, 15
   if hoje >= 15:
      min, max = 15, 31

   for celula in range(1, ws.max_row):
      numero = str(numero_dos_pagantes[celula].value)
      dia, month, ano = dia_vencimento(ws['E'], celula)

      if min < dia <= max and month == agora.month and ano == agora.year:
         numero_pagante = get_numero_formatado(numero)
         vencimento = get_vencimento(ws['E'], celula)
         numeros_e_vencimentos[numero_pagante] = vencimento

   return numeros_e_vencimentos

def get_numero_formatado(number: str) -> str:
   numero = number

   if is_brasileiro(number):
      numero = '55'+number

   return numero
