#!/usr/bin/env python3

from openpyxl.styles import Font
from openpyxl import Workbook, load_workbook

from functions.add_pagante import dados_to_planilha

def abrir_planilha(arquivo = '', mes = ''):
   try:
      wb = load_workbook(arquivo)
      ws = wb[mes]  

   # se não existir o arquivo, ele é criado
   except FileNotFoundError:
      wb = Workbook()
      ws = wb.active
      ws.title = mes       # renomea para o mês corrente.

      return quebra_galho(arquivo, wb, ws, mes, 'Arquivo não encontrado... Criando um...')

   # se não existir uma planilha com o nome do mês atual, esta é criada
   except KeyError:
      wb = load_workbook(arquivo)
      ws = wb.create_sheet(mes)
        
      return quebra_galho(arquivo, wb, ws, mes, "Planilha não encontrada... Criando uma...")
      
   else:
      return wb, ws

def quebra_galho(arquivo, wb, ws, mes, msg):
   dados = ('Nome', 'Número', 'E-mail', 'Dia de Pagamento', 'Vencimento')
   ft = Font(name='Arial', size=12, bold=True)
   alinhamento = ['center'] * 5

   print(msg)

   dados_to_planilha.insert_dados(1, ws, dados, alinhamento, ft)
   wb.save(arquivo)

   input("feito!!! Para continuar aperte Enter: ")
   return abrir_planilha(arquivo, mes)