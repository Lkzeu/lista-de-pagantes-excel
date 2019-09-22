#!/usr/bin/env python3

from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.keys import Keys

import os
import time
import pyperclip
import add_pagantes
import openpyxl as xl

def main():
   now = datetime.now()
   meses = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
             "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]         
   mes = meses[now.month - 1]

   chrome = "chromedriver.exe" if os.name == 'nt' else "chromedriver"

   while True:
      x = int(input("1 - adicionar pessoas\n"
                    "2 - enviar mensagens\n"))

      if x == 1:
         add_pagantes.cria_pagantes(mes)
      elif x == 2:
         enviar(chrome, mes)
      else:
         print("opção inválida!")
         continue
      break

def enviar(chromedrive = '', mes = ''):
   arquivo = "planilha/pagantes 2018.xlsx"
   wb = xl.load_workbook(arquivo)
   ws = wb.get_sheet_by_name(mes)
   vencimentos = ws['E']

    
   msg1 = "(Mensagem automática). Seu vencimento é dia "
   msg2 = ". Você tem até 1 dia após o vencimento para me enviar o comprovante, ou seu acesso será revogado."

   chromedriver = os.getcwd()+'/'+chromedrive
   driver = webdriver.Chrome(chromedriver)
   driver.get("https://web.whatsapp.com")
   time.sleep(20)

   while True:
      try:
         with open("TXTs/numeros55.txt", 'r') as number:
            i = 1
            for _ in number:
               _ = _.rstrip()
               venc = str(vencimentos[i].value)
            
               link = "https://web.whatsapp.com/send?phone="+_+"&amp;source=&amp;data="
               msg_completa = msg1 + venc + msg2
               pyperclip.copy(msg_completa)
            
               driver.get(link)
               time.sleep(10)
               input_box = driver.find_element_by_xpath('//*[@id="main"]/footer/div[1]/div[2]/div/div[2]') 
               #time.sleep(2)  
               input_box.send_keys(Keys.CONTROL+"v")
               time.sleep(2)
               input_box.send_keys(Keys.ENTER)
               time.sleep(10)

               i += 1
         break

      except IOError:
         rename(arquivo)
         continue

def rename(arquivo = ''): 
   wb = xl.load_workbook(arquivo)#read_only=True)
   ws = wb.active
   n_pagantes = ws['B']
   #vencimentos = ws['E']

   for celula in range(1, ws.max_row):
      numero = str(n_pagantes[celula].value)
      salva(numero)

def salva(number = ''):
   with open("TXTs/numeros55.txt", 'a+') as num:
      if is_brazil(str(number)):
         num.write("55"+number+'\n')
      else:
         num.write(number+'\n')

def is_brazil(telefone = ''):
   if len(telefone) == 10 or (len(telefone) == 11 and telefone[2] == '9'):
      with open("TXTs/ddd.txt", 'r') as a:
         ddds = a.read()
         if telefone[0:2] in ddds:
            return True
         else:
            return False
   return False

main()