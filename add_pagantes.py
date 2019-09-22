#!/usr/bin/env python3

from openpyxl.styles import Font, Alignment
from openpyxl import Workbook, load_workbook
from datetime import datetime
from shutil import copy2

import openpyxl
import os

def cria_pagantes(mes = ''):
   backup() # faz backup do arquivo

   arquivo = "planilha/pagantes 2018.xlsx"    
   try:
      wb = load_workbook(arquivo)
      ws = wb.get_sheet_by_name(mes)   

   # se não existir o arquivo, ele é criado
   except FileNotFoundError:
      wb = Workbook()
      ws = wb.active
      ws.title = mes       # renomea para o mês corrente.

      print("Arquivo não encontrado... Criando um...")
        
      id_columns(ws)
      wb.save(arquivo)

      input("feito!!! Para continuar aperte Enter: ")
      cria_pagantes(mes)

   # se não existir uma planilha com o nome do mês atual, esta é criada
   except KeyError:
      wb = load_workbook(arquivo)
      ws = wb.create_sheet(mes)
        
      print("Planilha não encontrada... Criando uma...")

      id_columns(ws)
      wb.save(arquivo)

      input("feito!!! Para continuar aperte Enter: ")
      cria_pagantes(mes)

   

   cl = 1  # a coluna deve ser 1 para inserir o Nome na coluna A.  NÃO MEXA, VALOR CONSTANTE!
   ln = ws.max_row + 1 # próxima linha vazia na planilha. Não sobreescreve linhas existentes.

   alinhamento = [Alignment(horizontal='center'), Alignment(horizontal='center'),
                  Alignment(horizontal='left'), Alignment(horizontal='right'), Alignment(horizontal='right')]
   ls_data = None # armazena os dados da pessoa 
   qt_meses = None # padrão de meses para o vencimento 

   dados = ('nome', 'numero', 'email', 'dia de pagamento [dd/mm]')  
   while True:
      nothing = os.system('cls' if os.name == 'nt' else 'clear') # limpa tela
      print("-"*40+"\nDados da {}ª pessoa!".format(ln-1))
      print("crtl + c para finalizar o programa!\n"+"-"*40)
      print("obs.: Dia de pagamento; aparte Enter para hoje ou s para 2 meses+\n\n")   

      while True:
         ls_data = [] # sempre que os dados estivem incorretos a lista é zerada
         qt_meses = 1 # "" o padrão é reestabelicido para 1

         try:
            # pega todos os dados das pessoas e armazena em uma lista vazia
            for _ in dados:
               x = input("Digite o {} da pessoa: ".format(_))
               if x == '':
                  x = hoje()
               if x.lower() == 's':
                  x = input("Qual o dia do pagamento? ")
                  qt_meses = int(input("Quantos meses {} tem? ".format(ls_data[0])))
               ls_data.append(x)
                
         except KeyboardInterrupt:
            print("\nPrograma finalizado!\n")
            wb.save(arquivo)
            quit(0)

         # caso tenha errado em algum dado, poderá inseri-los
         # novamente antes de colocá-los diretamente na planilha
         confirma = input("Todos os dados estão certos? [s/n] ")
         if confirma.lower() in ('sim'):
            break


      venc_day = vencimento(ls_data[3], qt_meses)  # index 3 é o dia do pagamento.
      ls_data.append(venc_day)
            
      # os dados são inseridos na planilha
      for _ in range(5):
         p = ws.cell(row=ln, column=(cl + _))
         p.value = ls_data[_]
         p.alignment = alinhamento[_]

      ln += 1   
      nothing = os.system('cls' if os.name == 'nt' else 'clear') # limpa tela
      
def hoje():
   now = datetime.now()
   dia, mes = now.day, now.month
   
   if dia < 10:
      dia = ''.join(['0', str(dia)])
   if mes < 10:
      mes = ''.join(['0', str(mes)])

   return ('/'.join([str(dia), str(mes)]))

def vencimento(data_pagamento = '00/00', meses = 0):
   dia = data_pagamento[:2]
   mes = data_pagamento[3:]
   ano = ''

   new_mes = int(mes) + meses
    
   if new_mes > 12:
      new_mes = new_mes - 12
      ano = ''.join(['/', '2020'])
   if new_mes < 10:
      new_mes = '0'+str(new_mes)

   return dia + '/' + str(new_mes) + ano

# cria a primeira linha que informa os dados de cada coluna, isto é Nome, Número...
def id_columns(ws):
   dados = ('Nome', 'Número', 'E-mail', 'Dia de Pagamento', 'Vencimento')
   ft = Font(name='Arial', size=12, bold=True)
   centro = Alignment(horizontal='center')
        
   for i in range(5):
      d = ws.cell(row=1, column=(1 + i))
      d.value = dados[i]
      d.font, d.alignment = ft, centro

def backup(modo = 0):
   now = datetime.now()
   data = '-'.join([str(now.day), str(now.month), str(now.year)])
   tempo = ':'.join([str(now.hour), str(now.minute), str(now.second)])
   versao = ' '.join([data, tempo])

   arquivo = os.getcwd()+"/planilha/pagantes 2018.xlsx"
   pasta_backup = os.getcwd() + "/backups/"
   destino = ''.join([pasta_backup, "pagantes backup", versao,".xlsx"])

   # Se algum parametro foi passado para a função.
   if modo:  # Número com valor zero é tratado como False em python
      os.system('cls' if os.name == 'nt' else 'clear') # limpa tela
      print("=-"*30)
      print("Vejo que é a primeira vez que executou o programa...\n"+"-"*60)
      print("Farei backup da sua planilha antes de fazer qualquer alteração")
      print("a fim de previnir a perda todos os dados por causa de algum bug de execução...")
      print("=-"*30)
      print("\nQuantos backups você quer manter armazenados? (os mais antigos serão excluídos) ")
      numero_backups = int(input("Digite a quantidade em número: "))

      with open("n_backup.dat", 'w') as bck:
         bck.write(str(numero_backups))
   else:
      try:
         with open("n_backup.dat", 'r') as r_bck:
            num_backups = r_bck.readline()
            backups_existentes = sorted(os.listdir(pasta_backup))

            if len(backups_existentes) == int(num_backups):
               os.remove(pasta_backup+backups_existentes[0])
      except IOError:
         backup(1)

   copy2(arquivo, destino)

